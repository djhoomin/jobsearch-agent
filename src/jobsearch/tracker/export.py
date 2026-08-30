"""Export the SQLite tracker to the user's existing spreadsheet shape.

The columns below were read off the real ``job-search-tracker.xlsx``:

* **Pipeline** - Company, Tier, Location, Stage / Round, Domain, IND sponsor,
  Buyer (20%), Role fit (25%), Company (25%), Domain fit (15%), Talent (15%),
  Weighted, Status, Warm path, Next action, Due, Notes
* **Outreach** - Name, Company / context, Relationship (A/B/C), Channel,
  Ask (intro/intel/referral), Sent, Follow-up due, Replied?, Outcome,
  'Who else?' referrals, Notes
* **Interviews** - Company, Role, Stage, Date, Interviewer(s),
  Prep notes / evidence to bring, Outcome, Next step
* **Dashboard** - Metric, Count

The user's own file is **never** written to. Exports go to the repo's output
directory, so a bad run can never destroy hand-maintained tracker rows.
"""

from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .db import Tracker

PIPELINE_COLUMNS = [
    "Company",
    "Tier",
    "Location",
    "Stage / Round",
    "Domain",
    "IND sponsor",
    "Buyer (20%)",
    "Role fit (25%)",
    "Company (25%)",
    "Domain fit (15%)",
    "Talent (15%)",
    "Weighted",
    "Status",
    "Warm path",
    "Next action",
    "Due",
    "Notes",
]

OUTREACH_COLUMNS = [
    "Name",
    "Company / context",
    "Relationship (A/B/C)",
    "Channel",
    "Ask (intro/intel/referral)",
    "Sent",
    "Follow-up due",
    "Replied?",
    "Outcome",
    "'Who else?' referrals",
    "Notes",
]

INTERVIEW_COLUMNS = [
    "Company",
    "Role",
    "Stage",
    "Date",
    "Interviewer(s)",
    "Prep notes / evidence to bring",
    "Outcome",
    "Next step",
]

DASHBOARD_COLUMNS = ["Metric", "Count"]

#: Extra sheet this tool adds. Kept separate so the user's four sheets keep
#: exactly the shape they already have.
JOBS_COLUMNS = [
    "Job ID",
    "Company",
    "Title",
    "Location",
    "Source",
    "URL",
    "Discovered",
    "Status",
    "Weighted",
    "Recommendation",
    "Constraints OK",
    "CV PDF",
    "Notes",
]

HEADER_FILL = PatternFill("solid", fgColor="1F2A37")
HEADER_FONT = Font(color="FFFFFF", bold=True)


class ExportError(RuntimeError):
    """Refused or impossible export."""


@dataclass
class ExportResult:
    path: Path
    rows: dict[str, int]

    def describe(self) -> str:
        counts = ", ".join(f"{sheet}: {n}" for sheet, n in self.rows.items())
        return f"{self.path} ({counts})"


def read_existing_columns(xlsx_path: str | Path) -> dict[str, list[str]]:
    """Read the header row of each sheet in the user's tracker.

    Used to keep the export compatible when the user adds a column: we follow
    their header row rather than our hardcoded list where the two disagree.
    """
    from openpyxl import load_workbook

    path = Path(xlsx_path)
    if not path.is_file():
        return {}
    workbook = load_workbook(str(path), read_only=True, data_only=True)
    shapes: dict[str, list[str]] = {}
    try:
        for name in workbook.sheetnames:
            sheet = workbook[name]
            header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            shapes[name] = [str(c) if c is not None else "" for c in header]
    finally:
        workbook.close()
    return shapes


def _status_of(row: sqlite3.Row) -> str:
    return row["status"] or ""


def _round(value: Any) -> Any:
    return round(value, 2) if isinstance(value, (int, float)) else value


def _pipeline_row(row: sqlite3.Row) -> list[Any]:
    return [
        row["company"],
        row["board_tier"],
        row["location"],
        "",  # Stage / Round: funding stage, maintained by hand
        row["department"] or "",
        {"yes": "YES", "no": "NO"}.get((row["ind_sponsor"] or "").lower(), "Pending"),
        _round(row["score_buyer"]),
        _round(row["score_role_fit"]),
        _round(row["score_company"]),
        _round(row["score_domain"]),
        _round(row["score_talent"]),
        _round(row["score_weighted"]),
        _status_of(row),
        row["warm_path"] or "",
        row["next_action"] or "",
        row["due"] or "",
        row["notes"] or "",
    ]


def _jobs_row(row: sqlite3.Row) -> list[Any]:
    return [
        row["job_id"],
        row["company"],
        row["title"],
        row["location"],
        row["source"],
        row["url"],
        (row["discovered_at"] or "")[:10],
        _status_of(row),
        _round(row["score_weighted"]),
        row["recommendation"] or "",
        {1: "yes", 0: "NO"}.get(row["constraints_ok"], "unchecked"),
        row["cv_pdf_path"] or "",
        (row["notes"] or "").replace("\n", " / ")[:500],
    ]


def _write_sheet(workbook: Workbook, title: str, columns: Sequence[str], rows: Sequence[Sequence[Any]]) -> int:
    sheet = workbook.create_sheet(title=title)
    sheet.append(list(columns))
    for index in range(1, len(columns) + 1):
        cell = sheet.cell(row=1, column=index)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in rows:
        sheet.append(list(row))
    sheet.freeze_panes = "A2"
    for index, name in enumerate(columns, start=1):
        width = 14 if len(name) < 12 else min(42, len(name) + 8)
        if name in {"Notes", "Next action", "Prep notes / evidence to bring", "URL"}:
            width = 48
        sheet.column_dimensions[get_column_letter(index)].width = width
    return len(rows)


def export_xlsx(
    tracker: Tracker,
    out_path: str | Path,
    *,
    template_xlsx: str | Path | None = None,
    protect_path: str | Path | None = None,
) -> ExportResult:
    """Write the tracker to a new xlsx in the user's column shape.

    ``protect_path`` is the user's own tracker: if ``out_path`` resolves to it,
    the export is refused outright.
    """
    out_path = Path(out_path)
    if protect_path is not None:
        protected = Path(protect_path).resolve()
        if out_path.resolve() == protected:
            raise ExportError(
                f"Refusing to overwrite your own tracker at {protected}. "
                "Exports are written to the repo's output directory instead."
            )
    out_path.parent.mkdir(parents=True, exist_ok=True)

    shapes = read_existing_columns(template_xlsx) if template_xlsx else {}
    pipeline_columns = shapes.get("Pipeline") or PIPELINE_COLUMNS
    outreach_columns = shapes.get("Outreach") or OUTREACH_COLUMNS
    interview_columns = shapes.get("Interviews") or INTERVIEW_COLUMNS
    dashboard_columns = shapes.get("Dashboard") or DASHBOARD_COLUMNS

    jobs = tracker.list_jobs()

    workbook = Workbook()
    workbook.remove(workbook.active)

    rows: dict[str, int] = {}

    pipeline_rows = [_fit(_pipeline_row(j), pipeline_columns, PIPELINE_COLUMNS) for j in jobs]
    rows["Pipeline"] = _write_sheet(workbook, "Pipeline", pipeline_columns, pipeline_rows)

    outreach_rows: list[list[Any]] = []
    for job in jobs:
        draft = tracker.latest_outreach(job["job_id"])
        for contact in tracker.contacts(job["job_id"]):
            outreach_rows.append(
                _fit(
                    [
                        contact["name"] or f"[{contact['title']}]",
                        f"{job['company']} - {job['title']}",
                        "C",  # cold by default; the user upgrades this by hand
                        "LinkedIn / email",
                        "intro",
                        "",
                        "",
                        "",
                        "",
                        "",
                        " | ".join(
                            filter(
                                None,
                                [
                                    contact["rationale"] or "",
                                    contact["search_url"] or "",
                                    f"draft ready: {draft['email_subject']}" if draft else "",
                                ],
                            )
                        )[:500],
                    ],
                    outreach_columns,
                    OUTREACH_COLUMNS,
                )
            )
    rows["Outreach"] = _write_sheet(workbook, "Outreach", outreach_columns, outreach_rows)

    interview_rows = [
        _fit(
            [job["company"], job["title"], job["status"], "", "", "", "", job["next_action"] or ""],
            interview_columns,
            INTERVIEW_COLUMNS,
        )
        for job in jobs
        if job["status"] in {"Interviewing", "In conversation"}
    ]
    rows["Interviews"] = _write_sheet(workbook, "Interviews", interview_columns, interview_rows)

    stats = tracker.stats()
    dashboard_rows = [
        ["Companies in pipeline", stats["total"]],
        ["Scored", stats["scored"]],
        ["Active (outreach through interviewing)", stats["active"]],
        ["Average weighted score", stats["avg_score"]],
        ["Best weighted score", stats["max_score"]],
        *[[f"Status: {status}", count] for status, count in sorted(stats["by_status"].items())],
        ["Exported", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
    ]
    rows["Dashboard"] = _write_sheet(
        workbook, "Dashboard", dashboard_columns, [_pad(r, len(dashboard_columns)) for r in dashboard_rows]
    )

    rows["Jobs (jobsearch-agent)"] = _write_sheet(
        workbook, "Jobs (jobsearch-agent)", JOBS_COLUMNS, [_jobs_row(j) for j in jobs]
    )

    workbook.save(str(out_path))
    return ExportResult(path=out_path, rows=rows)


def _fit(values: Sequence[Any], target: Sequence[str], canonical: Sequence[str]) -> list[Any]:
    """Reorder a row built for ``canonical`` into the user's ``target`` header.

    A column the user added that this tool knows nothing about comes out blank
    rather than shifting every subsequent value one cell to the left.
    """
    if list(target) == list(canonical):
        return list(values)
    lookup = {name: value for name, value in zip(canonical, values)}
    return [lookup.get(name, "") for name in target]


def _pad(values: Sequence[Any], width: int) -> list[Any]:
    padded = list(values)[:width]
    return padded + [""] * (width - len(padded))


__all__ = [
    "ExportError",
    "ExportResult",
    "PIPELINE_COLUMNS",
    "export_xlsx",
    "read_existing_columns",
]
