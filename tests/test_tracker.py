"""Tracker: state transitions, persistence, and the spreadsheet export."""

from __future__ import annotations

import pytest
from openpyxl import load_workbook

from jobsearch.models import JobPosting, Status, can_transition
from jobsearch.models import ConstraintReport, ConstraintResult, DimensionScore, ScoreReport, Verdict
from jobsearch.models import Contact, OutreachDraft
from jobsearch.tracker import Tracker, TrackerError
from jobsearch.tracker.export import (
    OUTREACH_COLUMNS,
    PIPELINE_COLUMNS,
    ExportError,
    export_xlsx,
    read_existing_columns,
)


@pytest.fixture
def tracker(tmp_path):
    with Tracker(path=tmp_path / "test.db") as t:
        yield t


@pytest.fixture
def job(tracker, posting):
    tracker.upsert_job(posting)
    return posting.job_id


# ---------------------------------------------------------------------------
# Transitions
# ---------------------------------------------------------------------------


class TestTransitionRules:
    def test_forward_progression_is_allowed(self):
        assert can_transition(Status.NOT_STARTED, Status.APPLIED)
        assert can_transition(Status.APPLIED, Status.INTERVIEWING)
        assert can_transition(Status.INTERVIEWING, Status.OFFER)

    def test_cannot_be_rejected_before_engaging(self):
        assert not can_transition(Status.NOT_STARTED, Status.REJECTED)

    def test_cannot_skip_from_rejected_back_into_process(self):
        assert not can_transition(Status.REJECTED, Status.INTERVIEWING)

    def test_rejected_can_only_be_reopened_as_parked(self):
        assert can_transition(Status.REJECTED, Status.PARKED)

    def test_offer_cannot_go_back_to_interviewing(self):
        assert not can_transition(Status.OFFER, Status.INTERVIEWING)

    def test_same_status_is_a_no_op_not_an_error(self):
        assert can_transition(Status.APPLIED, Status.APPLIED)

    def test_status_parsing_is_forgiving(self):
        assert Status.parse("applied") is Status.APPLIED
        assert Status.parse("In conversation") is Status.IN_CONVERSATION
        assert Status.parse(Status.OFFER) is Status.OFFER

    def test_unknown_status_raises_with_the_valid_set(self):
        with pytest.raises(ValueError, match="Valid:"):
            Status.parse("Ghosted")


class TestTrackerTransitions:
    def test_new_job_starts_not_started(self, tracker, job):
        assert tracker.get_job(job)["status"] == Status.NOT_STARTED.value

    def test_legal_transition_is_applied_and_logged(self, tracker, job):
        tracker.set_status(job, Status.APPLIED, reason="applied via careers page")
        assert tracker.get_job(job)["status"] == "Applied"
        history = tracker.history(job)
        assert history[-1]["from_status"] == "Not started"
        assert history[-1]["to_status"] == "Applied"
        assert history[-1]["reason"] == "applied via careers page"

    def test_illegal_transition_raises_and_changes_nothing(self, tracker, job):
        tracker.set_status(job, Status.APPLIED)
        tracker.set_status(job, Status.REJECTED)
        with pytest.raises(TrackerError, match="Illegal status transition"):
            tracker.set_status(job, Status.INTERVIEWING)
        assert tracker.get_job(job)["status"] == "Rejected"

    def test_error_message_lists_the_legal_moves(self, tracker, job):
        tracker.set_status(job, Status.APPLIED)
        tracker.set_status(job, Status.REJECTED)
        with pytest.raises(TrackerError, match="Parked"):
            tracker.set_status(job, Status.OFFER)

    def test_history_records_discovery(self, tracker, job):
        assert tracker.history(job)[0]["reason"] == "discovered"

    def test_repeat_status_does_not_duplicate_history(self, tracker, job):
        tracker.set_status(job, Status.APPLIED)
        before = len(tracker.history(job))
        tracker.set_status(job, Status.APPLIED)
        assert len(tracker.history(job)) == before

    def test_unknown_job_raises(self, tracker):
        with pytest.raises(TrackerError):
            tracker.set_status("nope", Status.APPLIED)


# ---------------------------------------------------------------------------
# Persistence
# ---------------------------------------------------------------------------


class TestPersistence:
    def test_upsert_is_idempotent(self, tracker, posting):
        tracker.upsert_job(posting)
        tracker.upsert_job(posting)
        assert len(tracker.list_jobs()) == 1

    def test_rediscovery_does_not_clobber_status_or_score(self, tracker, posting):
        tracker.upsert_job(posting)
        tracker.set_status(posting.job_id, Status.APPLIED)
        tracker.save_score(_score(posting.job_id, 4.25))

        posting.description = "Updated description from a fresh board fetch"
        tracker.upsert_job(posting)

        row = tracker.get_job(posting.job_id)
        assert row["status"] == "Applied"
        assert row["score_weighted"] == 4.25
        assert row["description"] == "Updated description from a fresh board fetch"

    def test_round_trips_to_a_posting(self, tracker, posting):
        tracker.upsert_job(posting)
        restored = tracker.get_posting(posting.job_id)
        assert restored.company == posting.company
        assert restored.title == posting.title
        assert restored.description == posting.description

    def test_resolve_accepts_a_prefix(self, tracker, job):
        assert tracker.resolve_job_id(job[:12]) == job

    def test_resolve_accepts_a_url(self, tracker, posting):
        tracker.upsert_job(posting)
        assert tracker.resolve_job_id(posting.url) == posting.job_id

    def test_resolve_rejects_an_unknown_id(self, tracker):
        with pytest.raises(TrackerError, match="No tracked job"):
            tracker.resolve_job_id("zzz")

    def test_ambiguous_prefix_raises(self, tracker):
        for i in range(2):
            tracker.upsert_job(
                JobPosting(company="Acme", title=f"Head of AI {i}", url=f"https://x/{i}",
                           job_id=f"acme-head-{i}")
            )
        with pytest.raises(TrackerError, match="ambiguous"):
            tracker.resolve_job_id("acme-head")

    def test_score_breakdown_is_persisted(self, tracker, job):
        tracker.save_score(_score(job, 4.4))
        row = tracker.get_job(job)
        assert row["score_weighted"] == 4.4
        assert row["score_buyer"] == 5
        assert row["constraints_ok"] == 1

    def test_outreach_and_contacts_persist(self, tracker, job):
        draft = OutreachDraft(
            job_id=job,
            contacts=[Contact(title="CTO", name="Ada", rationale="owns the mandate",
                              linkedin_search_url="https://www.linkedin.com/search/x")],
            linkedin_connection_note="note",
            linkedin_message="message",
            email_subject="subject",
            email_body="body",
        )
        tracker.save_outreach(draft)
        assert tracker.latest_outreach(job)["email_subject"] == "subject"
        assert tracker.contacts(job)[0]["name"] == "Ada"
        assert tracker.latest_outreach(job)["sent"] == 0, "nothing is ever marked sent by the tool"

    def test_saving_contacts_twice_replaces_rather_than_duplicates(self, tracker, job):
        tracker.save_contacts(job, [Contact(title="CTO")])
        tracker.save_contacts(job, [Contact(title="VP Engineering")])
        assert [c["title"] for c in tracker.contacts(job)] == ["VP Engineering"]

    def test_set_fields_rejects_unknown_columns(self, tracker, job):
        with pytest.raises(TrackerError, match="unknown field"):
            tracker.set_fields(job, status="Applied")

    def test_listing_filters(self, tracker, posting):
        tracker.upsert_job(posting)
        tracker.save_score(_score(posting.job_id, 4.4))
        tracker.upsert_job(JobPosting(company="Other", title="Head of AI", url="https://y/1"))
        assert len(tracker.list_jobs(min_score=4.0)) == 1
        assert len(tracker.list_jobs(company="other")) == 1
        assert len(tracker.list_jobs(status=Status.NOT_STARTED)) == 2

    def test_stats(self, tracker, posting):
        tracker.upsert_job(posting)
        tracker.save_score(_score(posting.job_id, 4.4))
        tracker.set_status(posting.job_id, Status.APPLIED)
        stats = tracker.stats()
        assert stats["total"] == 1
        assert stats["scored"] == 1
        assert stats["active"] == 1
        assert stats["max_score"] == 4.4


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------


class TestExport:
    def test_writes_the_users_four_sheets_plus_one(self, tracker, posting, tmp_path):
        tracker.upsert_job(posting)
        tracker.save_score(_score(posting.job_id, 4.4))
        out = tmp_path / "export.xlsx"
        export_xlsx(tracker, out)

        workbook = load_workbook(out)
        assert workbook.sheetnames == [
            "Pipeline",
            "Outreach",
            "Interviews",
            "Dashboard",
            "Jobs (jobsearch-agent)",
        ]

    def test_pipeline_header_matches_the_real_tracker(self, tracker, tmp_path):
        out = tmp_path / "export.xlsx"
        export_xlsx(tracker, out)
        header = [c.value for c in load_workbook(out)["Pipeline"][1]]
        assert header == PIPELINE_COLUMNS
        assert "Buyer (20%)" in header and "Weighted" in header

    def test_scores_land_in_the_right_columns(self, tracker, posting, tmp_path):
        tracker.upsert_job(posting)
        tracker.save_score(_score(posting.job_id, 4.4))
        out = tmp_path / "export.xlsx"
        export_xlsx(tracker, out)
        sheet = load_workbook(out)["Pipeline"]
        row = {sheet.cell(1, i).value: sheet.cell(2, i).value for i in range(1, sheet.max_column + 1)}
        assert row["Company"] == "Weaviate"
        assert row["Buyer (20%)"] == 5
        assert row["Weighted"] == 4.4
        assert row["IND sponsor"] == "YES"

    def test_refuses_to_overwrite_the_users_own_tracker(self, tracker, tmp_path):
        protected = tmp_path / "job-search-tracker.xlsx"
        protected.write_text("not really xlsx", encoding="utf-8")
        with pytest.raises(ExportError, match="Refusing to overwrite"):
            export_xlsx(tracker, protected, protect_path=protected)
        assert protected.read_text(encoding="utf-8") == "not really xlsx"

    def test_follows_a_users_extended_header(self, tracker, posting, tmp_path):
        """A column the user added must not shift every other value sideways."""
        template = tmp_path / "template.xlsx"
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Pipeline"
        sheet.append(["Company", "My Custom Column", "Weighted", "Status"])
        workbook.save(template)

        tracker.upsert_job(posting)
        tracker.save_score(_score(posting.job_id, 4.4))
        out = tmp_path / "export.xlsx"
        export_xlsx(tracker, out, template_xlsx=template)

        sheet = load_workbook(out)["Pipeline"]
        assert [c.value for c in sheet[1]] == ["Company", "My Custom Column", "Weighted", "Status"]
        # openpyxl stores an empty string as None; the point is that the
        # unknown column stays empty and does not shift Weighted or Status.
        assert [c.value for c in sheet[2]] == ["Weaviate", None, 4.4, "Not started"]

    def test_outreach_sheet_carries_contacts(self, tracker, posting, tmp_path):
        tracker.upsert_job(posting)
        tracker.save_outreach(
            OutreachDraft(
                job_id=posting.job_id,
                contacts=[Contact(title="CTO", name="Ada", linkedin_search_url="https://x")],
                email_subject="subject",
            )
        )
        out = tmp_path / "export.xlsx"
        export_xlsx(tracker, out)
        sheet = load_workbook(out)["Outreach"]
        assert [c.value for c in sheet[1]] == OUTREACH_COLUMNS
        assert sheet.cell(2, 1).value == "Ada"

    def test_interviews_sheet_only_lists_live_processes(self, tracker, posting, tmp_path):
        tracker.upsert_job(posting)
        out = tmp_path / "a.xlsx"
        export_xlsx(tracker, out)
        assert load_workbook(out)["Interviews"].max_row == 1  # header only

        tracker.set_status(posting.job_id, Status.APPLIED)
        tracker.set_status(posting.job_id, Status.INTERVIEWING)
        out2 = tmp_path / "b.xlsx"
        export_xlsx(tracker, out2)
        assert load_workbook(out2)["Interviews"].max_row == 2

    def test_reading_a_missing_template_is_not_fatal(self, tmp_path):
        assert read_existing_columns(tmp_path / "nope.xlsx") == {}


def _score(job_id: str, weighted: float) -> ScoreReport:
    constraints = ConstraintReport(
        results=[ConstraintResult("visa", Verdict.PASS, "known sponsor")]
    )
    return ScoreReport(
        job_id=job_id,
        constraints=constraints,
        dimensions=[
            DimensionScore("buyer", 5, 0.20),
            DimensionScore("role_fit", 5, 0.25),
            DimensionScore("company", 3.5, 0.25),
            DimensionScore("domain", 5, 0.15),
            DimensionScore("talent", 3.5, 0.15),
        ],
        weighted=weighted,
        recommendation="apply",
    )
